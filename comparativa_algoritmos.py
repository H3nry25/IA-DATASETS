# -*- coding: utf-8 -*-
"""
Comparativa de todos los algoritmos de ML para los 3 datasets:
  - Cáncer de Mama (Cancer de Mama)
  - Diabetes
  - Titanic

Algoritmos evaluados por dataset:
  1. Regresión Logística
  2. K-NN
  3. Decision Tree
  4. Random Forest

Genera: comparativa_algoritmos.xlsx
"""

import os
import warnings
import numpy as np
import pandas as pd

from sklearn.datasets import load_breast_cancer
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler, StandardScaler
from sklearn.linear_model import LogisticRegression
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import (accuracy_score, recall_score, precision_score,
                             f1_score, roc_auc_score, r2_score)
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings('ignore')

_DIR = os.path.dirname(os.path.abspath(__file__))

# =============================================================================
# HELPER — calcula todas las métricas de un resultado
# =============================================================================
def calcular_metricas(y_test, y_pred, y_pred_proba=None):
    metricas = {}
    metricas['Accuracy']    = round(accuracy_score(y_test, y_pred), 4)
    metricas['Recall']      = round(recall_score(y_test, y_pred, zero_division=0), 4)
    metricas['Precision']   = round(precision_score(y_test, y_pred, zero_division=0), 4)
    metricas['Specificity'] = round(recall_score(y_test, y_pred, pos_label=0, zero_division=0), 4)
    metricas['F1 Score']    = round(f1_score(y_test, y_pred, zero_division=0), 4)
    metricas['R2']          = round(r2_score(y_test, y_pred), 4)
    try:
        proba = y_pred_proba if y_pred_proba is not None else y_pred
        metricas['AUC'] = round(roc_auc_score(y_test, proba), 4)
    except ValueError:
        metricas['AUC'] = 'N/A'
    return metricas

resultados = []   # lista de dicts con todos los resultados

# =============================================================================
# 1. CÁNCER DE MAMA
# =============================================================================
print("=" * 60)
print("PROCESANDO: Cáncer de Mama")
print("=" * 60)

dataset = load_breast_cancer()
X_cm = dataset.data
y_cm = 1 - dataset.target   # 1=maligno (positivo), 0=benigno

X_train_cm, X_test_cm, y_train_cm, y_test_cm = train_test_split(
    X_cm, y_cm, test_size=0.2, random_state=42, stratify=y_cm)

scaler_cm = MinMaxScaler(feature_range=(0, 1))
X_train_cm_s = scaler_cm.fit_transform(X_train_cm)
X_test_cm_s  = scaler_cm.transform(X_test_cm)

# --- 1.1 Regresión Logística ---
print("  > Regresión Logística...")
m = LogisticRegression(solver='saga', max_iter=500)
m.fit(X_train_cm_s, y_train_cm)
yp = (m.predict(X_test_cm_s) > 0.5).astype(int)
yp_proba = m.predict_proba(X_test_cm_s)[:, 1]
resultados.append({'Dataset': 'Cáncer de Mama', 'Algoritmo': 'Regresión Logística',
                   **calcular_metricas(y_test_cm, yp, yp_proba)})

# --- 1.2 K-NN ---
print("  > K-NN...")
m = KNeighborsClassifier(n_neighbors=3, p=2, weights='uniform')
m.fit(X_train_cm_s, y_train_cm)
yp = m.predict(X_test_cm_s)
yp_proba = m.predict_proba(X_test_cm_s)[:, 1]
resultados.append({'Dataset': 'Cáncer de Mama', 'Algoritmo': 'K-NN',
                   **calcular_metricas(y_test_cm, yp, yp_proba)})

# --- 1.3 Decision Tree ---
print("  > Decision Tree...")
m = DecisionTreeClassifier(max_depth=7, criterion='gini', random_state=42)
m.fit(X_train_cm_s, y_train_cm)
yp = (m.predict(X_test_cm_s) > 0.5).astype(int)
yp_proba = m.predict_proba(X_test_cm_s)[:, 1]
resultados.append({'Dataset': 'Cáncer de Mama', 'Algoritmo': 'Decision Tree',
                   **calcular_metricas(y_test_cm, yp, yp_proba)})

# --- 1.4 Random Forest ---
print("  > Random Forest...")
m = RandomForestClassifier(n_estimators=15, max_depth=4, criterion='entropy', random_state=42)
m.fit(X_train_cm_s, y_train_cm)
yp = (m.predict(X_test_cm_s) > 0.5).astype(int)
yp_proba = m.predict_proba(X_test_cm_s)[:, 1]
resultados.append({'Dataset': 'Cáncer de Mama', 'Algoritmo': 'Random Forest',
                   **calcular_metricas(y_test_cm, yp, yp_proba)})

# =============================================================================
# 2. DIABETES
# =============================================================================
print("\n" + "=" * 60)
print("PROCESANDO: Diabetes")
print("=" * 60)

csv_diabetes = os.path.normpath(
    os.path.join(_DIR, 'Diabetes', '1. Preprocesamiento', 'diabetes_preprocessed.csv'))

try:
    df_diab = pd.read_csv(csv_diabetes)
    X_diab = df_diab.drop(columns=['Hubo_Hipoglucemia']).values
    y_diab = df_diab['Hubo_Hipoglucemia'].values

    X_train_d, X_test_d, y_train_d, y_test_d = train_test_split(
        X_diab, y_diab, test_size=0.2, random_state=42, stratify=y_diab)

    scaler_d = MinMaxScaler(feature_range=(0, 1))
    X_train_d_s = scaler_d.fit_transform(X_train_d)
    X_test_d_s  = scaler_d.transform(X_test_d)

    # --- 2.1 Regresión Logística ---
    print("  > Regresión Logística...")
    m = LogisticRegression(class_weight='balanced', solver='saga', max_iter=1000)
    m.fit(X_train_d_s, y_train_d)
    yp = (m.predict(X_test_d_s) > 0.5).astype(int)
    yp_proba = m.predict_proba(X_test_d_s)[:, 1]
    resultados.append({'Dataset': 'Diabetes', 'Algoritmo': 'Regresión Logística',
                       **calcular_metricas(y_test_d, yp, yp_proba)})

    # --- 2.2 K-NN ---
    print("  > K-NN...")
    m = KNeighborsClassifier(n_neighbors=3, p=2, weights='uniform')
    m.fit(X_train_d_s, y_train_d)
    yp = m.predict(X_test_d_s)
    yp_proba = m.predict_proba(X_test_d_s)[:, 1]
    resultados.append({'Dataset': 'Diabetes', 'Algoritmo': 'K-NN',
                       **calcular_metricas(y_test_d, yp, yp_proba)})

    # --- 2.3 Decision Tree ---
    print("  > Decision Tree...")
    m = DecisionTreeClassifier(max_depth=7, criterion='gini', class_weight='balanced', random_state=42)
    m.fit(X_train_d_s, y_train_d)
    yp = (m.predict(X_test_d_s) > 0.5).astype(int)
    yp_proba = m.predict_proba(X_test_d_s)[:, 1]
    resultados.append({'Dataset': 'Diabetes', 'Algoritmo': 'Decision Tree',
                       **calcular_metricas(y_test_d, yp, yp_proba)})

    # --- 2.4 Random Forest ---
    print("  > Random Forest...")
    m = RandomForestClassifier(n_estimators=15, max_depth=4, criterion='entropy',
                               class_weight='balanced', random_state=42)
    m.fit(X_train_d_s, y_train_d)
    yp = (m.predict(X_test_d_s) > 0.5).astype(int)
    yp_proba = m.predict_proba(X_test_d_s)[:, 1]
    resultados.append({'Dataset': 'Diabetes', 'Algoritmo': 'Random Forest',
                       **calcular_metricas(y_test_d, yp, yp_proba)})

except FileNotFoundError:
    print(f"  [!] No se encontró el CSV de Diabetes: {csv_diabetes}")
    print("      Ejecuta 'preprocess_data.py' del dataset Diabetes primero.")
    for alg in ['Regresión Logística', 'K-NN', 'Decision Tree', 'Random Forest']:
        resultados.append({'Dataset': 'Diabetes', 'Algoritmo': alg,
                           'Accuracy': 'N/A', 'Recall': 'N/A', 'Precision': 'N/A',
                           'Specificity': 'N/A', 'F1 Score': 'N/A', 'AUC': 'N/A', 'R2': 'N/A'})

# =============================================================================
# 3. TITANIC
# =============================================================================
print("\n" + "=" * 60)
print("PROCESANDO: Titanic")
print("=" * 60)

csv_titanic = os.path.normpath(
    os.path.join(_DIR, 'Titanic', 'Dataset', 'train_procesado_2.csv'))

try:
    df_tit = pd.read_csv(csv_titanic)
    X_tit = df_tit.drop(columns=['Survived']).values
    y_tit = df_tit['Survived'].values

    X_train_t, X_test_t, y_train_t, y_test_t = train_test_split(
        X_tit, y_tit, test_size=0.2, random_state=42, stratify=y_tit)

    # --- 3.1 Regresión Logística ---
    print("  > Regresión Logística...")
    m = LogisticRegression(class_weight='balanced', solver='saga', max_iter=1000)
    m.fit(X_train_t, y_train_t)
    yp = m.predict(X_test_t)
    yp_proba = m.predict_proba(X_test_t)[:, 1]
    resultados.append({'Dataset': 'Titanic', 'Algoritmo': 'Regresión Logística',
                       **calcular_metricas(y_test_t, yp, yp_proba)})

    # --- 3.2 K-NN ---
    print("  > K-NN...")
    m = KNeighborsClassifier(n_neighbors=3, p=2, weights='uniform')
    m.fit(X_train_t, y_train_t)
    yp = m.predict(X_test_t)
    yp_proba = m.predict_proba(X_test_t)[:, 1]
    resultados.append({'Dataset': 'Titanic', 'Algoritmo': 'K-NN',
                       **calcular_metricas(y_test_t, yp, yp_proba)})

    # --- 3.3 Decision Tree ---
    print("  > Decision Tree...")
    m = DecisionTreeClassifier(max_depth=4, criterion='gini', random_state=42)
    m.fit(X_train_t, y_train_t)
    yp = m.predict(X_test_t)
    yp_proba = m.predict_proba(X_test_t)[:, 1]
    resultados.append({'Dataset': 'Titanic', 'Algoritmo': 'Decision Tree',
                       **calcular_metricas(y_test_t, yp, yp_proba)})

    # --- 3.4 Random Forest ---
    print("  > Random Forest...")
    m = RandomForestClassifier(n_estimators=15, max_depth=4, criterion='entropy', random_state=42)
    m.fit(X_train_t, y_train_t)
    yp = m.predict(X_test_t)
    yp_proba = m.predict_proba(X_test_t)[:, 1]
    resultados.append({'Dataset': 'Titanic', 'Algoritmo': 'Random Forest',
                       **calcular_metricas(y_test_t, yp, yp_proba)})

except FileNotFoundError:
    print(f"  [!] No se encontró el CSV de Titanic: {csv_titanic}")
    for alg in ['Regresión Logística', 'K-NN', 'Decision Tree', 'Random Forest']:
        resultados.append({'Dataset': 'Titanic', 'Algoritmo': alg,
                           'Accuracy': 'N/A', 'Recall': 'N/A', 'Precision': 'N/A',
                           'Specificity': 'N/A', 'F1 Score': 'N/A', 'AUC': 'N/A', 'R2': 'N/A'})

# =============================================================================
# GENERAR EXCEL FORMATEADO
# =============================================================================
print("\n" + "=" * 60)
print("GENERANDO ARCHIVO EXCEL...")
print("=" * 60)

df_resultados = pd.DataFrame(resultados)
METRICAS = ['Accuracy', 'Recall', 'Precision', 'Specificity', 'F1 Score', 'AUC', 'R2']

# --- Colores de encabezado por dataset ---
COLORES_DATASET = {
    'Cáncer de Mama': {'header': '1565C0', 'sub': 'BBDEFB', 'alt': 'E3F2FD'},
    'Diabetes':       {'header': '2E7D32', 'sub': 'C8E6C9', 'alt': 'E8F5E9'},
    'Titanic':        {'header': '6A1B9A', 'sub': 'E1BEE7', 'alt': 'F3E5F5'},
}

COLOR_HEADER_MAIN = '263238'   # gris oscuro para la fila de títulos principales
COLOR_WHITE       = 'FFFFFF'
COLOR_METRIC_HDR  = '37474F'

wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 1: Comparativa Completa
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Comparativa Completa'

# Bordes
thin = Side(style='thin', color='B0BEC5')
thick = Side(style='medium', color='263238')
border_thin  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

def apply_border(cell, border):
    cell.border = border

# Título principal
ws.merge_cells('A1:I1')
title_cell = ws['A1']
title_cell.value = 'Comparativa de Algoritmos de Machine Learning — 3 Datasets'
title_cell.font = Font(bold=True, size=14, color=COLOR_WHITE, name='Calibri')
title_cell.fill = PatternFill('solid', fgColor=COLOR_HEADER_MAIN)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Subtítulo
ws.merge_cells('A2:I2')
sub_cell = ws['A2']
sub_cell.value = 'División: 80% entrenamiento / 20% prueba  |  random_state=42  |  Estratificado'
sub_cell.font = Font(italic=True, size=10, color='546E7A', name='Calibri')
sub_cell.fill = PatternFill('solid', fgColor='ECEFF1')
sub_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 18

# Encabezados de columna
headers = ['Dataset', 'Algoritmo'] + METRICAS
ws.row_dimensions[3].height = 22
for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font = Font(bold=True, size=11, color=COLOR_WHITE, name='Calibri')
    cell.fill = PatternFill('solid', fgColor=COLOR_METRIC_HDR)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin

# Anchos de columna
col_widths = [20, 22, 11, 11, 11, 13, 11, 11, 11]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Datos
ROW_START = 4
datasets_orden = ['Cáncer de Mama', 'Diabetes', 'Titanic']
fila = ROW_START

for ds in datasets_orden:
    df_ds = df_resultados[df_resultados['Dataset'] == ds]
    colores = COLORES_DATASET[ds]
    algoritmos = df_ds['Algoritmo'].tolist()

    # Fusionar celda de Dataset
    row_ini = fila
    row_fin = fila + len(algoritmos) - 1
    ws.merge_cells(f'A{row_ini}:A{row_fin}')
    ds_cell = ws[f'A{row_ini}']
    ds_cell.value = ds
    ds_cell.font = Font(bold=True, size=11, color=COLOR_WHITE, name='Calibri')
    ds_cell.fill = PatternFill('solid', fgColor=colores['header'])
    ds_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ds_cell.border = border_thick

    for i, row_data in enumerate(df_ds.itertuples(index=False)):
        bg = colores['sub'] if i % 2 == 0 else colores['alt']

        # Columna Algoritmo
        cell = ws.cell(row=fila, column=2, value=row_data.Algoritmo)
        cell.font = Font(bold=True, size=10, name='Calibri', color='212121')
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        cell.border = border_thin

        # Métricas numéricas
        for col_idx, metrica in enumerate(METRICAS, start=3):
            val = row_data._asdict().get(metrica)
            cell = ws.cell(row=fila, column=col_idx, value=val)
            cell.font = Font(size=10, name='Calibri')
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
            if isinstance(val, float):
                cell.number_format = '0.0000'

        ws.row_dimensions[fila].height = 18
        fila += 1

# Formato condicional (escala de color) para cada columna de métrica
for col_idx, metrica in enumerate(METRICAS, start=3):
    col_letter = get_column_letter(col_idx)
    rango = f'{col_letter}{ROW_START}:{col_letter}{fila-1}'
    # Verde = mejor, Rojo = peor (excepto R2 que puede ser negativo)
    ws.conditional_formatting.add(rango, ColorScaleRule(
        start_type='min', start_color='FF6B6B',
        mid_type='percentile', mid_value=50, mid_color='FFEB3B',
        end_type='max', end_color='66BB6A'))

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 2: Resumen por Algoritmo (Promedio entre datasets)
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Resumen por Algoritmo')

ws2.merge_cells('A1:I1')
c = ws2['A1']
c.value = 'Promedio de Métricas por Algoritmo (entre los 3 datasets)'
c.font = Font(bold=True, size=13, color=COLOR_WHITE, name='Calibri')
c.fill = PatternFill('solid', fgColor='4A148C')
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

headers2 = ['Algoritmo'] + METRICAS
for col_idx, h in enumerate(headers2, start=1):
    cell = ws2.cell(row=2, column=col_idx, value=h)
    cell.font = Font(bold=True, size=11, color=COLOR_WHITE, name='Calibri')
    cell.fill = PatternFill('solid', fgColor='6A1B9A')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin
ws2.row_dimensions[2].height = 20

df_num = df_resultados.copy()
for m in METRICAS:
    df_num[m] = pd.to_numeric(df_num[m], errors='coerce')

resumen = df_num.groupby('Algoritmo')[METRICAS].mean().round(4).reset_index()
ALG_COLORS = {
    'Regresión Logística': 'E8EAF6',
    'K-NN':                'FCE4EC',
    'Decision Tree':       'E8F5E9',
    'Random Forest':       'FFF8E1',
}
for i, row in resumen.iterrows():
    bg = ALG_COLORS.get(row['Algoritmo'], 'FFFFFF')
    for col_idx, col in enumerate(headers2, start=1):
        val = row[col]
        cell = ws2.cell(row=i+3, column=col_idx, value=val)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.border = border_thin
        cell.font = Font(size=10, name='Calibri',
                         bold=(col == 'Algoritmo'))
        cell.alignment = Alignment(horizontal='center' if col != 'Algoritmo' else 'left',
                                   vertical='center', indent=(1 if col == 'Algoritmo' else 0))
        if isinstance(val, float):
            cell.number_format = '0.0000'
    ws2.row_dimensions[i+3].height = 18

col_widths2 = [22] + [11]*7
for i, w in enumerate(col_widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Formato condicional hoja 2
for col_idx, metrica in enumerate(METRICAS, start=2):
    col_letter = get_column_letter(col_idx)
    rango = f'{col_letter}3:{col_letter}{3+len(resumen)-1}'
    ws2.conditional_formatting.add(rango, ColorScaleRule(
        start_type='min', start_color='FF6B6B',
        mid_type='percentile', mid_value=50, mid_color='FFEB3B',
        end_type='max', end_color='66BB6A'))

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 3: Resumen por Dataset
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet('Resumen por Dataset')

ws3.merge_cells('A1:I1')
c = ws3['A1']
c.value = 'Mejor Algoritmo por Dataset (según Accuracy)'
c.font = Font(bold=True, size=13, color=COLOR_WHITE, name='Calibri')
c.fill = PatternFill('solid', fgColor=COLOR_HEADER_MAIN)
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 28

headers3 = ['Dataset', 'Mejor Algoritmo'] + METRICAS
for col_idx, h in enumerate(headers3, start=1):
    cell = ws3.cell(row=2, column=col_idx, value=h)
    cell.font = Font(bold=True, size=11, color=COLOR_WHITE, name='Calibri')
    cell.fill = PatternFill('solid', fgColor=COLOR_METRIC_HDR)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin
ws3.row_dimensions[2].height = 20

for i, ds in enumerate(datasets_orden):
    colores = COLORES_DATASET[ds]
    df_ds_num = df_num[df_num['Dataset'] == ds].copy()
    if df_ds_num['Accuracy'].notna().any():
        best_idx = df_ds_num['Accuracy'].idxmax()
        best_row = df_ds_num.loc[best_idx]
    else:
        best_row = df_ds_num.iloc[0]

    row_excel = i + 3
    ws3.cell(row=row_excel, column=1, value=ds).font = Font(bold=True, size=10, color=COLOR_WHITE, name='Calibri')
    ws3.cell(row=row_excel, column=1).fill = PatternFill('solid', fgColor=colores['header'])
    ws3.cell(row=row_excel, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(row=row_excel, column=1).border = border_thin

    ws3.cell(row=row_excel, column=2, value=best_row['Algoritmo'])
    ws3.cell(row=row_excel, column=2).font = Font(bold=True, size=10, name='Calibri')
    ws3.cell(row=row_excel, column=2).fill = PatternFill('solid', fgColor=colores['sub'])
    ws3.cell(row=row_excel, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(row=row_excel, column=2).border = border_thin

    for col_idx, metrica in enumerate(METRICAS, start=3):
        val = best_row[metrica]
        cell = ws3.cell(row=row_excel, column=col_idx, value=val)
        cell.fill = PatternFill('solid', fgColor=colores['alt'])
        cell.font = Font(bold=True, size=10, name='Calibri', color='1B5E20' if isinstance(val, float) and val >= 0.9 else '212121')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
        if isinstance(val, float):
            cell.number_format = '0.0000'

    ws3.row_dimensions[row_excel].height = 22

col_widths3 = [20, 22] + [11]*7
for i, w in enumerate(col_widths3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ─────────────────────────────────────────────────────────────────────────────
# GUARDAR
# ─────────────────────────────────────────────────────────────────────────────
output_path = os.path.join(_DIR, 'comparativa_algoritmos.xlsx')
wb.save(output_path)

print(f"\n[OK] Archivo Excel generado exitosamente:")
print(f"   {output_path}")
print(f"\nContenido del Excel:")
print(f"  [Hoja 1] 'Comparativa Completa'  -- {len(resultados)} filas (3 datasets x 4 algoritmos)")
print(f"  [Hoja 2] 'Resumen por Algoritmo' -- promedio de metricas agrupado por algoritmo")
print(f"  [Hoja 3] 'Resumen por Dataset'   -- mejor algoritmo por dataset")
print(f"\nMetricas evaluadas: Accuracy, Recall, Precision, Specificity, F1 Score, AUC, R2")
